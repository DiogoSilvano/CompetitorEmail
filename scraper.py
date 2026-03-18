import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import pandas as pd
import time
import random
import re
import os
import io
from datetime import datetime, timedelta

HEADERS_LIST = [
    {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                      '(KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-GB,en;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
    },
    {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                      'AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15',
        'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
        'Accept-Language': 'en-GB,en;q=0.9',
        'Connection': 'keep-alive',
    }
]

SESSION = requests.Session()

# Shared Selenium driver — created once on first need, reused for all subsequent
# Selenium fallbacks, and shut down in main()'s finally block.
_shared_driver = None


# =============================================================================
# DATE / ID HELPERS
# =============================================================================

def get_week_ending(dt=None):
    """
    Return the EmailWeekEnding date string for a Friday-Thursday reporting week.
    The week is labelled by its closing Thursday.
    e.g. Fri 2026-02-20 through Thu 2026-02-26 -> "2026-02-26"
         Fri 2026-02-27 through Thu 2026-03-05 -> "2026-03-05"
    """
    if dt is None:
        dt = datetime.now()
    # Weekday: Mon=0 ... Thu=3, Fri=4 ... Sun=6
    # Days until next Thursday (inclusive of today if today is Thursday):
    # If Friday (4): +6 days to reach Thursday
    # If Saturday (5): +5 days
    # If Sunday (6): +4 days
    # If Monday (0): +3 days
    # If Tuesday (1): +2 days
    # If Wednesday (2): +1 day
    # If Thursday (3): +0 days
    days_to_thursday = (3 - dt.weekday()) % 7
    thursday = dt + timedelta(days=days_to_thursday)
    return thursday.strftime('%d/%m/%Y')


def get_scraped_date():
    """Return today as DD/MM/YYYY  00:00:00 matching Power Automate format."""
    return datetime.now().strftime('%d/%m/%Y  00:00:00')


# =============================================================================
# URL / TEXT UTILITIES
# =============================================================================

def get_source_name(url):
    try:
        hostname = urlparse(url).hostname or ''
        name = re.sub(r'^www\.', '', hostname)
        name = re.sub(r'\.(co\.uk|com|org|net|gov\.uk|org\.uk)$', '', name)
        name = name.replace('-', ' ').replace('.', ' ')
        return name.strip().title()
    except Exception:
        return ''


def normalise_url(url):
    """Strip query strings and fragments for URL-level duplicate detection."""
    try:
        p = urlparse(str(url))
        return f'{p.scheme}://{p.netloc}{p.path}'.rstrip('/')
    except Exception:
        return str(url).strip()


def is_binary_text(text):
    """Return True if text contains too many non-printable / binary characters."""
    if not text:
        return True
    sample = text[:500]
    bad = sum(1 for c in sample if ord(c) > 127 or (ord(c) < 32 and c not in '\n\r\t'))
    return (bad / len(sample)) > 0.1


def sanitise_str(value):
    """
    Remove XML 1.0 illegal characters that corrupt Excel / SharePoint files.
    These are the characters that cause the 'Repaired Records: String properties' error.
    Covers: null bytes, C0/C1 control chars, Unicode surrogates, non-characters.
    """
    if not isinstance(value, str):
        return value
    illegal = re.compile(
        r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F'   # control characters
        r'\uD800-\uDFFF'                        # UTF-16 surrogates
        r'\uFFFE\uFFFF]'                        # non-characters
    )
    return illegal.sub('', value).replace('\x00', '')


def sanitise_dataframe(df):
    """Apply sanitise_str to every string cell.
    Handles pandas >= 2.1 which renamed applymap to map."""
    result = df.copy()
    for col in result.select_dtypes(include='object').columns:
        result[col] = result[col].apply(
            lambda v: sanitise_str(v) if isinstance(v, str) else v)
    return result


# =============================================================================
# DATE NORMALISATION
# =============================================================================

def parse_any_date(value):
    """
    Try to parse a date string in any common format.
    Returns a datetime object or None.
    """
    if not isinstance(value, str):
        return None
    value = value.strip()
    if not value:
        return None
    formats = [
        '%d/%m/%Y',        # 26/02/2026
        '%d/%m/%Y  %H:%M:%S',  # 26/02/2026  00:00:00
        '%Y-%m-%d',        # 2026-02-26
        '%Y-%m-%d %H:%M:%S',
        '%m/%d/%Y',
        '%d-%m-%Y',
        '%B %d, %Y',
        '%d %B %Y',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def normalise_date_columns(df, columns=('DateScraped', 'EmailWeekEnding')):
    """
    Reformat all values in the given columns to dd/MM/yyyy.
    Values that cannot be parsed are left as-is.
    """
    df = df.copy()
    for col in columns:
        if col not in df.columns:
            continue
        def reformat(val):
            if pd.isna(val) or str(val).strip() == '':
                return val
            dt = parse_any_date(str(val))
            if dt:
                return dt.strftime('%d/%m/%Y')
            return val
        df[col] = df[col].apply(reformat)
    return df


# =============================================================================
# PDF HANDLING
# =============================================================================

def is_pdf_url(url):
    """Detect PDF links by URL extension or Content-Type header (HEAD request).
    HEAD request is only made when the extension is absent or ambiguous."""
    path = urlparse(str(url)).path.lower()
    if path.endswith('.pdf'):
        return True
    known_non_pdf = ('.html', '.htm', '.php', '.asp', '.aspx', '.xml', '.json', '.txt')
    if any(path.endswith(ext) for ext in known_non_pdf):
        return False
    # Extension absent or ambiguous — fall back to a HEAD request
    try:
        head = SESSION.head(url, timeout=10, allow_redirects=True)
        return 'pdf' in head.headers.get('Content-Type', '').lower()
    except Exception:
        return False


def scrape_pdf(url):
    """
    Download a PDF and extract clean text with pdfplumber (must be installed).
    Install: pip install pdfplumber

    Returns the same dict shape as scrape_article so callers need no special casing.
    """
    try:
        import pdfplumber
    except ImportError:
        return {
            'normalized_url': url, 'title': '', 'text': '',
            'source_name': get_source_name(url),
            'accessible': False, 'method': 'pdf',
            'error': 'pdfplumber not installed -- run: pip install pdfplumber',
        }

    try:
        headers = random.choice(HEADERS_LIST)
        last_err = None
        r = None
        for attempt in range(3):
            try:
                session = SESSION if attempt == 0 else requests.Session()
                r = session.get(url, headers=headers, timeout=30, allow_redirects=True)
                break
            except Exception as e:
                last_err = e
                time.sleep(2 ** attempt)
        if r is None:
            return {
                'normalized_url': url, 'title': '', 'text': '',
                'source_name': get_source_name(url),
                'accessible': False, 'method': 'pdf',
                'error': f'Connection failed after 3 attempts: {str(last_err)[:80]}',
            }
        if r.status_code != 200:
            return {
                'normalized_url': url, 'title': '', 'text': '',
                'source_name': get_source_name(url),
                'accessible': False, 'method': 'pdf',
                'error': f'HTTP {r.status_code}',
            }

        with pdfplumber.open(io.BytesIO(r.content)) as pdf:
            pages_text = []
            for page in pdf.pages[:15]:     # cap at 15 pages
                t = page.extract_text()
                if t:
                    pages_text.append(t)
            full_text = '\n'.join(pages_text).strip()

        if not full_text:
            return {
                'normalized_url': url, 'title': '', 'text': '',
                'source_name': get_source_name(url),
                'accessible': False, 'method': 'pdf',
                'error': 'PDF text extraction returned empty',
            }

        lines = [l.strip() for l in full_text.splitlines() if len(l.strip()) > 10]
        title = lines[0][:200] if lines else ''

        return {
            'normalized_url': r.url, 'title': title,
            'text': full_text[:8000],
            'source_name': get_source_name(r.url),
            'accessible': True, 'method': 'pdf', 'error': None,
        }

    except Exception as e:
        return {
            'normalized_url': url, 'title': '', 'text': '',
            'source_name': get_source_name(url),
            'accessible': False, 'method': 'pdf',
            'error': f'PDF error: {str(e)[:80]}',
        }


# =============================================================================
# HTML SCRAPING
# =============================================================================

def extract_title(soup):
    og = soup.find('meta', property='og:title')
    if og and og.get('content'):
        return og['content'].strip()
    if soup.title and soup.title.string:
        return soup.title.string.strip()
    h1 = soup.find('h1')
    if h1:
        return h1.get_text(strip=True)
    return ''


def extract_text(html):
    soup = BeautifulSoup(html, 'lxml')
    title = extract_title(soup)
    for tag in soup(['script', 'style', 'nav', 'footer', 'header',
                     'aside', 'form', 'iframe', 'noscript']):
        tag.decompose()
    body = (
        soup.find('article') or
        soup.find(class_=lambda c: c and any(x in c.lower() for x in
            ['article-body', 'post-content', 'entry-content',
             'article-content', 'story-body'])) or
        soup.find('main') or
        soup.find('body')
    )
    if not body:
        return title, ''
    paragraphs = body.find_all('p')
    text = ' '.join(
        p.get_text(separator=' ', strip=True)
        for p in paragraphs
        if len(p.get_text(strip=True)) > 40
        and 'cookie' not in p.get_text(strip=True).lower()
    )
    if len(text) < 200:
        text = body.get_text(separator=' ', strip=True)
    return title, text[:8000]


def scrape_with_requests(url):
    try:
        time.sleep(random.uniform(1.5, 3.5))
        headers = random.choice(HEADERS_LIST)
        r = SESSION.get(url, headers=headers, timeout=20, allow_redirects=True)
        final_url = r.url
        if r.status_code == 403: return final_url, '', '', '403'
        if r.status_code == 429: return final_url, '', '', '429'
        if r.status_code != 200: return final_url, '', '', f'HTTP {r.status_code}'
        if r.encoding and r.encoding.lower() == 'iso-8859-1':
            r.encoding = 'utf-8'
        title, text = extract_text(r.text)
        if is_binary_text(text): return final_url, '', '', 'binary_content'
        if len(text) < 100: return final_url, title, '', 'too_short'
        return final_url, title, text, None
    except Exception as e:
        return url, '', '', f'Connection:{str(e)[:60]}'


def _get_or_create_driver():
    """Return the shared Selenium driver, creating it on first call."""
    global _shared_driver
    if _shared_driver is None:
        options = Options()
        options.add_argument('--headless=new')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument(
            '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'
        )
        service = Service(ChromeDriverManager().install())
        _shared_driver = webdriver.Chrome(service=service, options=options)
        _shared_driver.set_page_load_timeout(30)
    return _shared_driver


def _shutdown_driver():
    """Quit and discard the shared Selenium driver if it exists."""
    global _shared_driver
    if _shared_driver is not None:
        try:
            _shared_driver.quit()
        except Exception:
            pass
        _shared_driver = None


def scrape_with_selenium(url, driver=None):
    """Scrape a URL using Selenium.

    If *driver* is supplied the caller owns its lifecycle (no quit on exit).
    If *driver* is None a temporary single-use driver is created and quit after use.
    """
    owns_driver = driver is None
    if owns_driver:
        driver = _get_or_create_driver()

    try:
        driver.get(url)
        time.sleep(random.uniform(3, 5))
        final_url = driver.current_url
        for by, selector in [
            (By.ID, 'accept-cookies'),
            (By.ID, 'onetrust-accept-btn-handler'),
            (By.CSS_SELECTOR, '.cookie-accept'),
            (By.XPATH, "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
                       "'abcdefghijklmnopqrstuvwxyz'),'accept all')]"),
            (By.XPATH, "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
                       "'abcdefghijklmnopqrstuvwxyz'),'accept cookies')]"),
            (By.XPATH, "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
                       "'abcdefghijklmnopqrstuvwxyz'),'agree')]"),
        ]:
            try:
                btn = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((by, selector)))
                btn.click()
                time.sleep(2)
                break
            except Exception:
                pass
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight / 2);")
        time.sleep(2)
        title = driver.execute_script("""
            const og = document.querySelector('meta[property="og:title"]');
            if (og) return og.content;
            const h1 = document.querySelector('h1');
            if (h1) return h1.innerText.trim();
            return document.title;
        """)
        text = driver.execute_script("""
            const selectors = ['article p', 'main p',
                '[class*="content"] p', '[class*="article"] p', 'p'];
            for (const sel of selectors) {
                const paras = Array.from(document.querySelectorAll(sel))
                    .map(p => p.innerText.trim())
                    .filter(t => t.length > 40 && !t.toLowerCase().includes('cookie'));
                if (paras.length > 3) return paras.slice(0, 100).join(' ');
            }
            return '';
        """)
        text = (text or '')[:8000]
        if len(text) < 100:
            return final_url, title or '', '', 'too_short after browser render'
        return final_url, title or '', text, None
    except Exception as e:
        return url, '', '', f'Selenium error: {str(e)[:80]}'
    finally:
        if owns_driver:
            try:
                driver.quit()
            except Exception:
                pass


def scrape_article(url):
    """
    Route URL to the correct scraper:
    - PDFs    -> pdfplumber extractor
    - HTML    -> requests, falling back to Selenium

    Always returns dict with keys:
        normalized_url, title, text, source_name, accessible, method, error
    """
    if is_pdf_url(url):
        print(f'   [PDF] Detected — extracting with pdfplumber')
        return scrape_pdf(url)

    final_url, title, text, err = scrape_with_requests(url)
    if text:
        return {
            'normalized_url': final_url, 'title': title, 'text': text,
            'source_name': get_source_name(final_url),
            'accessible': True, 'method': 'requests', 'error': None,
        }
    if any(x in str(err) for x in ('403', '429', 'too_short', 'Connection', 'binary_content')):
        print(f'   [WARN] requests failed ({err[:40]}), trying Selenium...')
        final_url2, title2, text2, err2 = scrape_with_selenium(
            url, driver=_get_or_create_driver())
        if text2:
            return {
                'normalized_url': final_url2, 'title': title2, 'text': text2,
                'source_name': get_source_name(final_url2),
                'accessible': True, 'method': 'selenium', 'error': None,
            }
        return {
            'normalized_url': final_url2, 'title': '', 'text': '',
            'source_name': get_source_name(final_url2),
            'accessible': False, 'method': 'selenium', 'error': err2,
        }
    return {
        'normalized_url': final_url, 'title': '', 'text': '',
        'source_name': get_source_name(final_url),
        'accessible': False, 'method': 'requests', 'error': err,
    }


# =============================================================================
# EXCEL WRITER
# =============================================================================

def rewrite_excel_table(output_path, df):
    """
    Write df to the named Excel table 'ArticlesTable' in sheet 'Articles'.
    Strategy: write a clean new sheet via ExcelWriter (which never corrupts),
    then open with openpyxl only to attach/rename the table definition.
    This avoids the delete-rows + append pattern that desynchronises table refs.
    """
    SHEET_NAME  = 'Articles'
    TABLE_NAME  = 'ArticlesTable'
    TABLE_STYLE = 'TableStyleMedium9'

    import tempfile, shutil

    # ── Step 1: write clean data to a temp file ──────────────────────────────
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(tmp_fd)

    try:
        with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=SHEET_NAME)

        # ── Step 2: open temp file and add/update the table definition ────────
        wb = load_workbook(tmp_path)
        ws = wb[SHEET_NAME]

        n_rows = len(df)
        n_cols = len(df.columns)
        ref = f'A1:{get_column_letter(n_cols)}{n_rows + 1}'

        # Remove any auto-generated table openpyxl may have added
        for tname in list(ws.tables.keys()):
            del ws.tables[tname]

        tab = Table(displayName=TABLE_NAME, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name=TABLE_STYLE, showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)

        wb.save(tmp_path)
        wb.close()

        # ── Step 3: atomically replace the output file ────────────────────────
        shutil.move(tmp_path, output_path)
        print(f'Written {n_rows} rows to {os.path.basename(output_path)}')

    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise

# =============================================================================
# MAIN
# =============================================================================

def main():
    DATA_DIR   = r'C:\Users\dsn24\OneDrive - Sky\Diogo\Competitor Email Automation\Python Script\data'
    OUTPUT_DIR = r'C:\Users\dsn24\OneDrive - Sky\Diogo\Competitor Email Automation\Python Script\output'

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    input_file  = os.path.join(DATA_DIR,   'Competitor_Email_DB.xlsx')
    output_path = os.path.join(OUTPUT_DIR, 'Competitor_Email_DB.xlsx')

    if not os.path.exists(input_file):
        print(f'ERROR: File not found: {input_file}')
        input('Press Enter to exit...')
        return

    # =========================================================================
    # STEP 1: Scrape pending email links
    # =========================================================================
    print('=' * 55)
    print('STEP 1: Scraping email newsletter links')
    print('=' * 55)

    # Always read Pending rows from the input file (Competitor_Email_DB.xlsx in /data).
    # Power Automate writes new Pending rows there.
    print(f'Reading from: {os.path.basename(input_file)}')
    try:
        df = pd.read_excel(input_file, sheet_name='Articles', engine='openpyxl')
    except Exception:
        df = pd.read_excel(input_file, engine='openpyxl')

    for col in ['NormalizedURL', 'Title', 'Full_Article_Text',
                'SourceName', 'Status', 'EmailWeekEnding', 'DateScraped']:
        if col not in df.columns:
            df[col] = ''
        df[col] = df[col].astype(object)

    week_ending  = get_week_ending()
    scraped_date = get_scraped_date()

    print(f'Total rows  : {len(df)}')
    print(f'EmailWeekEnding : {week_ending}')
    print(f'DateScraped : {scraped_date}')

    # Diagnostic: show what Status values exist in the file
    status_counts = df['Status'].value_counts(dropna=False).to_dict()
    print(f'Status breakdown: {status_counts}')

    # Strip whitespace from Status — Power Automate sometimes adds trailing spaces
    df['Status'] = df['Status'].astype(str).str.strip()

    pending = df[df['Status'] == 'Pending']
    success_email = failed_email = 0
    print(f'Pending rows: {len(pending)}\n')

    total_pending = len(pending)
    try:
        for pending_num, (i, row) in enumerate(pending.iterrows(), start=1):
            url = str(row.get('ArticleURL', '')).strip()
            if not url or url.lower() == 'nan':
                df.at[i, 'Status'] = 'Skipped'
                continue

            print(f'[{pending_num}/{total_pending}] {url[:90]}')
            result = scrape_article(url)

            if result['accessible']:
                df.at[i, 'NormalizedURL']     = result['normalized_url']
                df.at[i, 'Title']             = result['title']
                df.at[i, 'Full_Article_Text'] = result['text']
                df.at[i, 'SourceName']        = result['source_name']
                df.at[i, 'Status']            = 'Analysed'
                df.at[i, 'EmailWeekEnding']   = week_ending
                existing_ds = row.get('DateScraped', '')
                if pd.isna(existing_ds) or str(existing_ds).strip() == '':
                    df.at[i, 'DateScraped'] = scraped_date
                print(f'   OK [{result["method"]}] {str(result["title"])[:70]}')
                success_email += 1
            else:
                df.at[i, 'NormalizedURL'] = result['normalized_url']
                df.at[i, 'SourceName']    = result['source_name']
                df.at[i, 'Status']        = 'Failed'
                print(f'   FAIL: {result["error"]}')
                failed_email += 1
    finally:
        _shutdown_driver()

    print(f'\nEmail scrape done -- {success_email} success, {failed_email} failed\n')

    # =========================================================================
    # STEP 2: Build final output and save
    # =========================================================================
    print('=' * 55)
    print('STEP 2: Writing output')
    print('=' * 55)

    output_df = df[df['Status'] == 'Analysed'].copy().reset_index(drop=True)

    if 'EmailWeekEnding' not in output_df.columns:
        output_df['EmailWeekEnding'] = ''

    # Normalise date columns to dd/MM/yyyy before writing
    output_df = normalise_date_columns(output_df)

    # Sanitise all strings before writing -- prevents Excel / SharePoint corruption
    output_df = sanitise_dataframe(output_df)
    rewrite_excel_table(output_path, output_df)

    # =========================================================================
    # SUMMARY
    # =========================================================================
    print('\n' + '=' * 55)
    print('SUMMARY')
    print('=' * 55)
    print(f'Email scraped       : {success_email} OK, {failed_email} failed')
    print(f'Total rows in file  : {len(output_df)}')
    print(f'Output              : {output_path}')

    input('\nPress Enter to exit...')


if __name__ == '__main__':
    main()
