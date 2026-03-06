import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import pandas as pd
import time
import random
import re
import os
import hashlib
from datetime import datetime, timedelta
import feedparser

HEADERS_LIST = [
    {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-GB,en;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
    },
    {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15',
        'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
        'Accept-Language': 'en-GB,en;q=0.9',
        'Connection': 'keep-alive',
    }
]

SESSION = requests.Session()


def generate_link_id(url):
    return hashlib.md5(url.encode()).hexdigest()[:12].upper()


def get_current_week_friday():
    """Return the Friday that started the current weekly window (most recent past Friday)."""
    today = datetime.now()
    # weekday(): Monday=0, Friday=4
    days_since_friday = (today.weekday() - 4) % 7
    friday = today - timedelta(days=days_since_friday)
    return friday.replace(hour=0, minute=0, second=0, microsecond=0)


def get_last_friday():
    """Get the date of the previous Friday (start of current weekly window)."""
    return get_current_week_friday()


def get_week_number():
    return datetime.now().strftime('%W')


def get_week_ending(dt=None):
    """Return closing Thursday of the Fri-Thu reporting week as YYYY-MM-DD."""
    if dt is None:
        dt = datetime.now()
    days_to_thursday = (3 - dt.weekday()) % 7
    thursday = dt + timedelta(days=days_to_thursday)
    return thursday.strftime('%Y-%m-%d')
def get_weekly_filename(output_dir):
    """
    Return the filename for this week's RSS file.
    Filename is anchored to the Friday that started the week, e.g.:
      RSS_Articles_W09_2026_from20260227.xlsx
    """
    friday = get_current_week_friday()
    week_num = friday.strftime('%W')
    year = friday.strftime('%Y')
    date_str = friday.strftime('%Y%m%d')
    return os.path.join(output_dir, f'RSS_Articles_W{week_num}_{year}_from{date_str}.xlsx')


def get_feed_name(feed_url):
    try:
        hostname = urlparse(feed_url).hostname or ''
        name = re.sub(r'^www\.', '', hostname)
        name = re.sub(r'\.(co\.uk|com|org|net|gov\.uk|org\.uk)$', '', name)
        return name.strip().title()
    except:
        return 'Unknown'


def get_source_name(url):
    try:
        hostname = urlparse(url).hostname or ''
        name = re.sub(r'^www\.', '', hostname)
        name = re.sub(r'\.(co\.uk|com|org|net|gov\.uk|org\.uk)$', '', name)
        name = name.replace('-', ' ').replace('.', ' ')
        return name.strip().title()
    except:
        return ''


def is_relevant_article(title, description):
    """
    Check relevance using title only. Tightened keyword list to reduce
    international/enterprise noise while preserving UK telco & streaming coverage.
    """
    text = title.lower()

    keywords = [
        # ── UK Broadband / Fixed ──────────────────────────────────────────
        'broadband', 'fibre', 'fttp', 'full fibre', 'gigabit',
        'openreach', 'cityfibre', 'altnet', 'hyperoptic', 'netomnia',
        'nexfibre', 'gofibre', 'project gigabit', 'bduk',
        'italk', 'zen internet', 'brsk',

        # ── UK Mobile operators (named, not generic '5g') ─────────────────
        'vodafone uk', 'vodafone three', 'vmo2', 'virgin media o2',
        'virgin media', 'sky mobile', 'sky broadband', 'sky glass', 'sky stream',
        'talktalk', 'three uk', 'o2 uk', 'o2 satellite', 'plusnet',
        'bt broadband', 'bt group', 'bt consumer', 'bt business',
        'ee broadband', 'ee mobile',

        # ── Satellite (UK context) ────────────────────────────────────────
        'starlink uk', 'direct-to-device', 'direct-to-cell',

        # ── Streaming services (business news — named specifically) ───────
        'netflix', 'disney+', 'prime video', 'paramount+', 'now tv',
        'dazn', 'hbo max', 'apple tv+', 'warner bros', 'wbd',
        'svod', 'avod',

        # ── Sports rights (business context) ─────────────────────────────
        'sports rights', 'broadcast rights', 'premier league rights',
        'champions league rights', 'world cup rights',

        # ── UK Broadcast & regulation ─────────────────────────────────────
        'ofcom', 'bbc ', 'itv ', 'channel 4', 'channel 5',
        'cma ruling', 'ofcom ruling', 'ofcom fine', 'ofcom report',
        'online safety act', 'age verification',

        # ── Piracy (UK/named context, not generic) ────────────────────────
        'sports piracy', 'iptv piracy', 'anti-piracy uk',
        'illegal streaming uk', 'bitplay', 'piracy action',

        # ── Business events ───────────────────────────────────────────────
        'merger', 'acquisition', 'takeover',
        'price hike', 'price rise', 'price increase', 'tariff',
        'quarterly results', 'annual results', 'subscriber',

        # ── Named stakeholders ────────────────────────────────────────────
        'giacom', 'e& vodafone', 'e&\'s stake in vodafone',
    ]

    return any(kw in text for kw in keywords)


def is_binary_text(text):
    if not text:
        return True
    non_printable = sum(1 for c in text[:500] if ord(c) > 127 or (ord(c) < 32 and c not in '\n\r\t'))
    return (non_printable / min(len(text), 500)) > 0.1


def extract_title(soup):
    og = soup.find('meta', property='og:title')
    if og and og.get('content'):
        return og['content'].strip()
    if soup.title and soup.title.string:
        return soup.title.string.strip()
    h1 = soup.find('h1')
    if h1:
        return h1.get_text(strip=True)
    return ''


def extract_text(html):
    soup = BeautifulSoup(html, 'lxml')
    title = extract_title(soup)
    for tag in soup(['script', 'style', 'nav', 'footer', 'header', 'aside', 'form', 'iframe', 'noscript']):
        tag.decompose()
    body = (
        soup.find('article') or
        soup.find(class_=lambda c: c and any(x in c.lower() for x in
            ['article-body', 'post-content', 'entry-content', 'article-content', 'story-body'])) or
        soup.find('main') or
        soup.find('body')
    )
    if not body:
        return title, ''
    paragraphs = body.find_all('p')
    text = ' '.join(
        p.get_text(separator=' ', strip=True)
        for p in paragraphs
        if len(p.get_text(strip=True)) > 40
        and 'cookie' not in p.get_text(strip=True).lower()
    )
    if len(text) < 200:
        text = body.get_text(separator=' ', strip=True)
    return title, text[:8000]


def scrape_with_requests(url):
    try:
        time.sleep(random.uniform(1.5, 3.5))
        headers = random.choice(HEADERS_LIST)
        r = SESSION.get(url, headers=headers, timeout=20, allow_redirects=True)
        final_url = r.url
        if r.status_code == 403: return final_url, '', '', '403'
        if r.status_code == 429: return final_url, '', '', '429'
        if r.status_code != 200: return final_url, '', '', f'HTTP {r.status_code}'
        if r.encoding and r.encoding.lower() == 'iso-8859-1':
            r.encoding = 'utf-8'
        html = r.text
        title, text = extract_text(html)
        if is_binary_text(text): return final_url, '', '', 'binary_content'
        if len(text) < 100: return final_url, title, '', 'too_short'
        return final_url, title, text, None
    except Exception as e:
        return url, '', '', f'Connection:{str(e)[:60]}'


def scrape_with_selenium(url):
    driver = None
    try:
        options = Options()
        options.add_argument('--headless=new')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_experimental_option('excludeSwitches', ['enable-automation'])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36')
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(30)
        driver.get(url)
        time.sleep(random.uniform(3, 5))
        final_url = driver.current_url
        for by, selector in [
            (By.ID, 'accept-cookies'),
            (By.ID, 'onetrust-accept-btn-handler'),
            (By.CSS_SELECTOR, '.cookie-accept'),
            (By.XPATH, "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'accept all')]"),
            (By.XPATH, "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'accept cookies')]"),
            (By.XPATH, "//button[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'agree')]"),
        ]:
            try:
                btn = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((by, selector)))
                btn.click()
                time.sleep(2)
                break
            except:
                pass
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight / 2);")
        time.sleep(2)
        title = driver.execute_script("""
            const og = document.querySelector('meta[property="og:title"]');
            if (og) return og.content;
            const h1 = document.querySelector('h1');
            if (h1) return h1.innerText.trim();
            return document.title;
        """)
        text = driver.execute_script("""
            const selectors = ['article p','main p','[class*="content"] p','[class*="article"] p','p'];
            for (const sel of selectors) {
                const paras = Array.from(document.querySelectorAll(sel))
                    .map(p => p.innerText.trim())
                    .filter(t => t.length > 40 && !t.toLowerCase().includes('cookie'));
                if (paras.length > 3) return paras.slice(0, 100).join(' ');
            }
            return '';
        """)
        text = text[:8000] if text else ''
        if len(text) < 100: return final_url, title or '', '', 'too_short after browser render'
        return final_url, title or '', text, None
    except Exception as e:
        return url, '', '', f'Selenium error: {str(e)[:80]}'
    finally:
        if driver: driver.quit()


def scrape_article(url):
    final_url, title, text, err = scrape_with_requests(url)
    if text:
        return {'normalized_url': final_url, 'title': title, 'text': text,
                'source_name': get_source_name(final_url), 'accessible': True, 'method': 'requests', 'error': None}
    if any(x in str(err) for x in ('403', '429', 'too_short', 'Connection', 'binary_content')):
        print(f'   ⚠️  requests failed ({err[:40]}), trying Selenium...')
        final_url2, title2, text2, err2 = scrape_with_selenium(url)
        if text2:
            return {'normalized_url': final_url2, 'title': title2, 'text': text2,
                    'source_name': get_source_name(final_url2), 'accessible': True, 'method': 'selenium', 'error': None}
        return {'normalized_url': final_url2, 'title': '', 'text': '',
                'source_name': get_source_name(final_url2), 'accessible': False, 'method': 'selenium', 'error': err2}
    return {'normalized_url': final_url, 'title': '', 'text': '',
            'source_name': get_source_name(final_url), 'accessible': False, 'method': 'requests', 'error': err}


def fetch_rss_articles(rss_feeds_file):
    """Fetch relevant articles from RSS feeds published since last Friday."""
    if not os.path.exists(rss_feeds_file):
        print(f'⚠️  RSS feeds file not found: {rss_feeds_file}')
        return []

    with open(rss_feeds_file, 'r') as f:
        feed_urls = [line.strip() for line in f if line.strip() and not line.startswith('#')]

    if not feed_urls:
        print('⚠️  No RSS feeds configured')
        return []

    last_friday = get_last_friday()
    articles = []

    print(f'\n📡 Fetching RSS articles since {last_friday.strftime("%Y-%m-%d")}...')

    for feed_url in feed_urls:
        try:
            print(f'   Fetching: {feed_url}')
            feed = feedparser.parse(feed_url)
            feed_name = get_feed_name(feed_url)

            relevant_count = 0
            for entry in feed.entries:
                pub_date = None
                if hasattr(entry, 'published_parsed') and entry.published_parsed:
                    pub_date = datetime(*entry.published_parsed[:6])
                elif hasattr(entry, 'updated_parsed') and entry.updated_parsed:
                    pub_date = datetime(*entry.updated_parsed[:6])

                if pub_date and pub_date >= last_friday:
                    article_url = entry.link if hasattr(entry, 'link') else None
                    title = entry.title if hasattr(entry, 'title') else ''
                    description = entry.description if hasattr(entry, 'description') else ''

                    if article_url and is_relevant_article(title, description):
                        articles.append({
                            'url': article_url,
                            'feed_name': feed_name,
                            'pub_date': pub_date
                        })
                        relevant_count += 1

            print(f'      ✓ {relevant_count} relevant articles')

        except Exception as e:
            print(f'   ❌ Error fetching {feed_url}: {str(e)[:60]}')

    print(f'   ✅ Total new from feeds: {len(articles)} relevant articles\n')
    return articles


def load_existing_weekly_file(output_path):
    """Load existing rows from this week's file if it exists."""
    if not os.path.exists(output_path):
        return pd.DataFrame(), set()

    try:
        df = pd.read_excel(output_path, sheet_name='RSS_Articles', engine='openpyxl')
        existing_urls = set(df['NormalizedURL'].dropna().astype(str).tolist())
        print(f'📂 Loaded existing weekly file: {len(df)} rows, {len(existing_urls)} known URLs')
        return df, existing_urls
    except Exception as e:
        print(f'⚠️  Could not read existing weekly file ({e}), starting fresh.')
        return pd.DataFrame(), set()


def save_weekly_file(output_path, df):
    """Write/overwrite the weekly file, preserving the named table."""
    SHEET_NAME = 'RSS_Articles'
    TABLE_NAME = 'RSSArticlesTable'
    TABLE_STYLE = 'TableStyleMedium9'

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=SHEET_NAME)
        ws = writer.sheets[SHEET_NAME]

        max_row = len(df) + 1
        max_col = len(df.columns)
        table_ref = f'A1:{get_column_letter(max_col)}{max_row}'

        tab = Table(displayName=TABLE_NAME, ref=table_ref)
        style = TableStyleInfo(
            name=TABLE_STYLE, showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

    print(f'📗 Saved weekly file: {len(df)} total rows → {output_path}')


def main():
    SCRIPT_DIR = r'C:\Users\dsn24\OneDrive - Sky\Diogo\Competitor Email Automation\Python Script\RSS Parser'
    OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'output')
    RSS_FILE   = os.path.join(SCRIPT_DIR, 'rss_feeds.txt')

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ── Determine this week's output file ────────────────────────────────────
    output_path = get_weekly_filename(OUTPUT_DIR)
    friday = get_current_week_friday()
    print(f'📅 Weekly window: {friday.strftime("%Y-%m-%d")} (Friday) → Thursday')
    print(f'📁 Weekly file  : {os.path.basename(output_path)}\n')

    # ── Load existing rows from this week's file (if any) ────────────────────
    existing_df, existing_urls = load_existing_weekly_file(output_path)

    # ── Fetch new RSS articles ────────────────────────────────────────────────
    rss_articles = fetch_rss_articles(RSS_FILE)

    if not rss_articles:
        print('ℹ️  No new RSS articles found')
        input('\nPress Enter to exit...')
        return

    # ── Filter out URLs already in this week's file ───────────────────────────
    new_articles = [a for a in rss_articles if a['url'] not in existing_urls]
    skipped = len(rss_articles) - len(new_articles)
    if skipped:
        print(f'⏭️  Skipped {skipped} articles already in this week\'s file\n')

    if not new_articles:
        print('ℹ️  All fetched articles already recorded this week.')
        input('\nPress Enter to exit...')
        return

    # ── Scrape new articles ───────────────────────────────────────────────────
    week_num  = get_week_number()
    today     = datetime.now().strftime('%Y-%m-%d')

    rows = []
    success = 0
    failed  = 0

    print(f'🔍 Scraping {len(new_articles)} new RSS articles...\n')

    for idx, article in enumerate(new_articles, 1):
        url       = article['url']
        feed_name = article['feed_name']

        print(f'[{idx}/{len(new_articles)}] {url[:80]}')
        result = scrape_article(url)

        if result['accessible']:
            link_id      = generate_link_id(result['normalized_url'])
            article_date = article['pub_date'].strftime('%Y-%m-%d')
            row = {
                'LinkID':            link_id,
                'ArticleURL':        '',
                'TrackingURL':       '',
                'NormalizedURL':     result['normalized_url'],
                'Title':             result['title'],
                'Full_Article_Text': result['text'],
                'SourceEmail':       f'RSS - {feed_name}',
                'SourceName':        result['source_name'],
                'Status':            'Analysed',
                'DateScraped':       article_date,
                'WeekNumber':        week_num,
                'EmailWeekEnding':   get_week_ending(),
            }
            rows.append(row)
            print(f'   ✅ [{result["method"]}] {result["title"][:70]}')
            success += 1
        else:
            print(f'   ❌ {result["error"]}')
            failed += 1

    if not rows:
        print('\n❌ No new articles successfully scraped')
        input('\nPress Enter to exit...')
        return

    # ── Combine with existing rows and save ───────────────────────────────────
    new_df     = pd.DataFrame(rows)
    combined   = pd.concat([existing_df, new_df], ignore_index=True) if not existing_df.empty else new_df
    save_weekly_file(output_path, combined)

    print('\n' + '━' * 45)
    print(f'✅ New scraped    : {success}')
    print(f'❌ Failed         : {failed}')
    print(f'📦 Total in file  : {len(combined)}')
    print(f'📁 File           : {os.path.basename(output_path)}')
    input('\nPress Enter to exit...')


if __name__ == '__main__':
    main()